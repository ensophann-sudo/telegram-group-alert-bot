import os
import sqlite3
from datetime import datetime, time, timedelta
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)


BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
TIMEZONE = ZoneInfo("Asia/Phnom_Penh")
DB_FILE = "telegram_group_messages.db"


def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_id TEXT,
        group_name TEXT,
        sender_name TEXT,
        sender_username TEXT,
        message_text TEXT,
        has_photo TEXT,
        message_id TEXT,
        message_link TEXT,
        date_text TEXT,
        time_text TEXT,
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS config (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    """)

    conn.commit()
    conn.close()


def set_config(key, value):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("""
    INSERT INTO config (key, value)
    VALUES (?, ?)
    ON CONFLICT(key) DO UPDATE SET value = excluded.value
    """, (key, value))

    conn.commit()
    conn.close()


def get_config(key):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("SELECT value FROM config WHERE key = ?", (key,))
    row = cur.fetchone()

    conn.close()

    return row[0] if row else None


def build_message_link(chat, message_id):
    try:
        if chat.username:
            return f"https://t.me/{chat.username}/{message_id}"

        chat_id = str(chat.id)

        if chat_id.startswith("-100"):
            internal_id = chat_id.replace("-100", "")
            return f"https://t.me/c/{internal_id}/{message_id}"

        return ""
    except Exception:
        return ""


def save_message(
    group_id,
    group_name,
    sender_name,
    sender_username,
    message_text,
    has_photo,
    message_id,
    message_link,
    date_text,
    time_text,
    created_at
):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("""
    INSERT INTO messages (
        group_id,
        group_name,
        sender_name,
        sender_username,
        message_text,
        has_photo,
        message_id,
        message_link,
        date_text,
        time_text,
        created_at
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        group_id,
        group_name,
        sender_name,
        sender_username,
        message_text,
        has_photo,
        message_id,
        message_link,
        date_text,
        time_text,
        created_at
    ))

    conn.commit()
    conn.close()


def get_messages_last_7_days():
    now = datetime.now(TIMEZONE)
    seven_days_ago = now - timedelta(days=7)

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("""
    SELECT
        group_name,
        sender_name,
        sender_username,
        message_text,
        has_photo,
        date_text,
        time_text,
        created_at
    FROM messages
    WHERE created_at >= ?
    ORDER BY created_at ASC
    """, (seven_days_ago.isoformat(),))

    rows = cur.fetchall()
    conn.close()

    return rows


def get_all_latest_messages():
    """
    If there are no messages in the last 7 days,
    send all saved latest data from database.
    """
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("""
    SELECT
        group_name,
        sender_name,
        sender_username,
        message_text,
        has_photo,
        date_text,
        time_text,
        created_at
    FROM messages
    ORDER BY created_at ASC
    """)

    rows = cur.fetchall()
    conn.close()

    return rows


def get_report_messages():
    """
    Main report logic:
    1. Try to get messages from last 7 days.
    2. If no messages in last 7 days, get all saved latest data.
    """
    rows = get_messages_last_7_days()

    if rows:
        return rows, "last 7 days"

    rows = get_all_latest_messages()

    if rows:
        return rows, "all saved latest data"

    return [], "no data"


def group_messages_by_same_datetime(rows):
    """
    Group by:
    Group + Sender + Username + Date + Time HH:MM:SS

    If text and photos are sent at the same date/time HH:MM:SS,
    they will show only one row in Excel.

    Photo column shows photo quantity.
    """
    grouped = {}

    for row in rows:
        (
            group_name,
            sender_name,
            sender_username,
            message_text,
            has_photo,
            date_text,
            time_text,
            created_at
        ) = row

        time_hhmmss = str(time_text)[:8]
        date_time_text = f"{date_text} {time_hhmmss}"

        key = (
            group_name or "",
            sender_name or "",
            sender_username or "",
            date_time_text
        )

        if key not in grouped:
            grouped[key] = {
                "group_name": group_name or "",
                "sender_name": sender_name or "",
                "sender_username": sender_username or "",
                "texts": [],
                "photo_count": 0,
                "date": date_time_text,
                "created_at": created_at or date_time_text
            }

        if message_text:
            grouped[key]["texts"].append(str(message_text))

        if has_photo == "Yes":
            grouped[key]["photo_count"] += 1

    return grouped


def create_excel_file(rows):
    now = datetime.now(TIMEZONE)
    file_name = f"telegram_group_messages_{now.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    grouped = group_messages_by_same_datetime(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Telegram Messages"

    # Correct table headers only
    headers = [
        "Name Group",
        "Name Sender",
        "Username",
        "Text",
        "Photo",
        "Date"
    ]

    ws.append(headers)

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sorted_items = sorted(
        grouped.values(),
        key=lambda x: x["created_at"]
    )

    for data in sorted_items:
        combined_text = "\n".join(data["texts"]) if data["texts"] else ""

        photo_count = data["photo_count"]

        if photo_count == 0:
            photo_text = "No"
        elif photo_count == 1:
            photo_text = "1 photo"
        else:
            photo_text = f"{photo_count} photos"

        ws.append([
            data["group_name"],
            data["sender_name"],
            data["sender_username"],
            combined_text,
            photo_text,
            data["date"]
        ])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    ws.freeze_panes = "A2"

    wb.save(file_name)
    return file_name


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat

    if chat.type == "private":
        await update.message.reply_text(
            f"Hello Sophann!\n\n"
            f"Your Chat ID: {chat.id}\n\n"
            f"Use /setme to save this chat for auto report.\n"
            f"Use /sendnow to receive Excel immediately."
        )
    else:
        await update.message.reply_text("Bot is active in this group.")


async def setme(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat

    if chat.type != "private":
        await update.message.reply_text("Please use /setme in private chat with the bot.")
        return

    set_config("personal_chat_id", str(chat.id))

    await update.message.reply_text(
        "Done! Your chat ID is saved.\n"
        "Auto Excel report will be sent every Friday at 4 PM Cambodia time."
    )


async def status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    personal_chat_id = get_config("personal_chat_id")
    last_7_rows = get_messages_last_7_days()
    all_rows = get_all_latest_messages()

    await update.message.reply_text(
        "Bot Status\n\n"
        f"Chat ID saved: {'Yes' if personal_chat_id else 'No'}\n"
        f"Messages in last 7 days: {len(last_7_rows)}\n"
        f"All saved messages: {len(all_rows)}\n"
        "Auto report: Every Friday at 4 PM Cambodia time\n"
        "Excel table: Name Group, Name Sender, Username, Text, Photo, Date"
    )


async def sendnow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    rows, report_type = get_report_messages()

    if not rows:
        await update.message.reply_text("No saved data in database yet.")
        return

    file_name = create_excel_file(rows)

    with open(file_name, "rb") as f:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=f,
            caption=(
                f"Excel report sent now.\n"
                f"Report data: {report_type}\n"
                f"Saved records used: {len(rows)}"
            )
        )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "/start - Show bot info\n"
        "/setme - Save your private chat ID for auto report\n"
        "/status - Check bot status\n"
        "/sendnow - Send Excel report now\n"
        "/help - Show help"
    )


async def collect_group_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message

    if not message:
        return

    chat = update.effective_chat
    user = update.effective_user
    now = datetime.now(TIMEZONE)

    group_id = str(chat.id)
    group_name = chat.title or "Unknown Group"

    if user:
        sender_name = user.full_name
        sender_username = f"@{user.username}" if user.username else ""
    else:
        sender_name = "Unknown Sender"
        sender_username = ""

    message_text = message.text or message.caption or ""
    has_photo = "Yes" if message.photo else "No"
    message_id = str(message.message_id)
    message_link = build_message_link(chat, message.message_id)

    date_text = now.strftime("%Y-%m-%d")
    time_text = now.strftime("%H:%M:%S")
    created_at = now.isoformat()

    save_message(
        group_id=group_id,
        group_name=group_name,
        sender_name=sender_name,
        sender_username=sender_username,
        message_text=message_text,
        has_photo=has_photo,
        message_id=message_id,
        message_link=message_link,
        date_text=date_text,
        time_text=time_text,
        created_at=created_at
    )

    print(
        f"Saved | Group: {group_name} | Sender: {sender_name} | "
        f"Photo: {has_photo} | Date: {date_text} {time_text}"
    )


async def send_weekly_excel(context: ContextTypes.DEFAULT_TYPE):
    personal_chat_id = get_config("personal_chat_id")
    now = datetime.now(TIMEZONE)

    if not personal_chat_id:
        print("No personal chat ID saved. Please send /setme to the bot in private chat.")
        return

    rows, report_type = get_report_messages()

    if not rows:
        await context.bot.send_message(
            chat_id=personal_chat_id,
            text=f"No saved data in database yet.\nChecked: {now.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        return

    file_name = create_excel_file(rows)

    with open(file_name, "rb") as f:
        await context.bot.send_document(
            chat_id=personal_chat_id,
            document=f,
            caption=(
                "Weekly Telegram Excel Report\n"
                f"Auto sent: {now.strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Report data: {report_type}\n"
                f"Saved records used: {len(rows)}"
            )
        )

    print(f"Weekly Excel report sent at {now.strftime('%Y-%m-%d %H:%M:%S')}")


def main():
    if not BOT_TOKEN:
        print("ERROR: TELEGRAM_BOT_TOKEN is not set.")
        print("Please set TELEGRAM_BOT_TOKEN environment variable.")
        return

    init_db()

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("setme", setme))
    app.add_handler(CommandHandler("status", status))
    app.add_handler(CommandHandler("sendnow", sendnow))
    app.add_handler(CommandHandler("help", help_command))

    app.add_handler(MessageHandler(
        filters.ChatType.GROUPS
        & (filters.TEXT | filters.PHOTO)
        & ~filters.COMMAND,
        collect_group_message
    ))

    # Auto send every Friday at 4:00 PM Cambodia time
    # Monday = 0, Tuesday = 1, Wednesday = 2, Thursday = 3, Friday = 4
    app.job_queue.run_daily(
        send_weekly_excel,
        time=time(hour=16, minute=0, second=0, tzinfo=TIMEZONE),
        days=(4,),
        name="weekly_excel_friday_4pm"
    )

    print("Bot is running...")
    print("Auto Excel report: Every Friday at 4 PM Cambodia time.")
    print("Use /sendnow to send Excel immediately.")
    print("If no data in last 7 days, bot will send all saved latest data.")
    print("Excel columns: Name Group, Name Sender, Username, Text, Photo, Date")
    print("Collecting group messages silently...")

    app.run_polling()


if __name__ == "__main__":
    main()
